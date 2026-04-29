from flask import Flask, jsonify, send_file
import pandas as pd
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google.oauth2 import service_account
import io, os, tempfile
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

FOLDER_ID = os.environ.get("FOLDER_ID")
CREDS_JSON = os.environ.get("GOOGLE_CREDENTIALS")

def get_drive_service():
    import json
    creds_info = json.loads(CREDS_JSON)
    creds = service_account.Credentials.from_service_account_info(
        creds_info,
        scopes=["https://www.googleapis.com/auth/drive"]
    )
    return build("drive", "v3", credentials=creds)

def download_file(service, file_id):
    request = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf

@app.route("/comparar", methods=["GET"])
def comparar():
    try:
        service = get_drive_service()

        results = service.files().list(
            q=f"'{FOLDER_ID}' in parents and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and trashed=false",
            fields="files(id, name, createdTime)",
            orderBy="createdTime desc"
        ).execute()

        files = results.get("files", [])

        if len(files) < 2:
            return jsonify({"error": "Se necesitan al menos 2 archivos Excel"}), 400

        buf1 = download_file(service, files[0]["id"])
        buf2 = download_file(service, files[1]["id"])

        liv = pd.read_excel(buf1)
        gym = pd.read_excel(buf2)

        # Normalizar
        liv["SKU_norm"] = liv.iloc[:, 0].astype(str).str.strip().str.upper()
        gym["SKU_norm"] = gym.iloc[:, 0].astype(str).str.strip().str.upper()

        liv = liv[(liv["SKU_norm"].notna()) & (liv["SKU_norm"] != "NAN") & (liv["SKU_norm"] != "")]
        gym = gym[(gym["SKU_norm"].notna()) & (gym["SKU_norm"] != "NAN") & (gym["SKU_norm"] != "")]

        liv["_qty"] = pd.to_numeric(liv.iloc[:, 1], errors="coerce").fillna(0)
        gym["_qty"] = pd.to_numeric(gym.iloc[:, 1], errors="coerce").fillna(0)

        liv_idx = liv.set_index("SKU_norm")
        gym_agg = gym.groupby("SKU_norm")["_qty"].sum()

        # SOLO productos en ambos
        en_ambos = set(liv["SKU_norm"]) & set(gym["SKU_norm"])

        rows = []

        for sku in sorted(en_ambos):
            q1 = float(liv_idx.loc[sku, "_qty"])
            q2 = float(gym_agg.get(sku, 0))
            diff = q2 - q1

            if diff > 0:
                accion = "Subir en Liverpool"
            elif diff < 0:
                accion = "Bajar en Liverpool"
            else:
                accion = "OK"

            rows.append({
                "SKU": sku,
                "Liverpool": q1,
                "Almacén": q2,
                "Diferencia": diff,
                "Acción": accion
            })

        df_result = pd.DataFrame(rows)

        # Crear Excel
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        df_result.to_excel(tmp.name, index=False)

        # 🎨 Colores
        wb = load_workbook(tmp.name)
        ws = wb.active

        rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        azul = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

        for row in range(2, ws.max_row + 1):
            accion = ws[f"E{row}"].value

            if accion == "Subir en Liverpool":
                fill = rojo
            elif accion == "Bajar en Liverpool":
                fill = azul
            else:
                fill = verde

            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = fill

        wb.save(tmp.name)

        return send_file(
            tmp.name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Comparacion_Inteligente.xlsx"
        )

    except Exception as e:
        import traceback
        return jsonify({
            "error": str(e),
            "detalle": traceback.format_exc()
        }), 500